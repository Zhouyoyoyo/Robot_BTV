from openpyxl import load_workbook


def load_excel_kv(path):
    wb = load_workbook(path)
    ws = wb.active

    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key, value = row
        data[key] = value
    return data
